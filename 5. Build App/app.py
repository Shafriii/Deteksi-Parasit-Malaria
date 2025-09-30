# -*- coding: utf-8 -*-
import os, sys, locale, datetime, csv, subprocess
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk
import numpy as np
import cv2
from PIL import Image, ImageTk
import onnxruntime as ort
# --- Simple tooltip helper (works for Tk/CustomTkinter) ---
def add_tooltip(widget, text: str):
    tip = tk.Toplevel(widget)
    tip.withdraw()
    tip.overrideredirect(True)
    tip.attributes("-topmost", True)

    # warna auto menyesuaikan mode
    try:
        is_light = ctk.get_appearance_mode().lower() == "light"
    except Exception:
        is_light = False
    bg = "#111827" if not is_light else "#111827"   # gelap tetap kontras
    fg = "#ffffff"

    lbl = tk.Label(tip, text=text, bg=bg, fg=fg, padx=6, pady=3, borderwidth=1, relief="solid")
    lbl.pack()

    def enter(e):
        x = e.x_root + 12
        y = e.y_root + 10
        tip.geometry(f"+{x}+{y}")
        tip.deiconify()

    def leave(_):
        tip.withdraw()

    widget.bind("<Enter>", enter)
    widget.bind("<Leave>", leave)

# Excel (optional)
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

# ────────────────────────── Language packs ──────────────────────────
LANG = {
    "en": {
        "window_title": "Malaria Parasite Detection",
        "left_panel": "Original",
        "right_panel": "Detections",
        "counts_title_img": "Counts — This Image",
        "counts_title_sess": "Counts — Session Total",
        "hdr_totals": "Grand totals",
        # Sidebar / menus
        "sidebar_title": "Controls",
        "menu_file": "File",
        "menu_session": "Session",
        "menu_view": "View",
        "menu_choose": "Choose…",
        "open_image": "Open Image",
        "process_folder": "Process Folder",
        "save_result": "Save Result (current image)",
        "save_session": "Save Session (all)",
        "add_to_session": "Add Current to Session",
        "reset_session": "Reset Session",
        "fit": "Fit to Window",
        "reset": "Reset View",
        "appearance_label": "Appearance",
        "appearance_system": "System",
        "appearance_light": "Light",
        "appearance_dark": "Dark",
        "lang_label": "Language",
        "session_images": "Session images:",
        "sidebar_toggle_tip": "Hide/Show Sidebar",
        "sidebar_expand": "Show",
        "sidebar_collapse": "Hide",
        # Status / dialogs
        "status_idle": "Open an image to begin. Scroll to zoom, drag to pan.",
        "loading_model": "Loading model…",
        "model_loaded": "Model loaded.",
        "weights_not_found_title": "Model not found",
        "weights_not_found_msg": "best.onnx not found in the bundle.",
        "choose_image_title": "Choose an image",
        "open_error_title": "Open Error",
        "open_error_msg": "Failed to read the image.",
        "no_image_title": "No Image",
        "no_image_msg": "Please open an image first.",
        "running": "Running inference…",
        "no_results": "No detections.",
        "detections_saved": "Detections saved:",
        "save_result_title": "Save Result",
        "png_filter": "PNG",
        "jpeg_filter": "JPEG",
        "all_files": "All files",
        "saved_prefix": "Saved:",
        "save_error_title": "Save Error",
        "save_error_msg": "Failed to write image.",
        "loaded_prefix": "Loaded:",
        "nothing_to_save": "Nothing to save yet.\nRun detection then add to Session.",
        "save_session_title": "Choose Session Folder",
        "session_saved": "Session saved to:",
        "add_before_detect": "Run detection first before adding to session.",
        "added_entry": "Added to session.",
        "reset_done": "Session reset.",
        "run_detection": "Run Detection",
        "step2_hint": "",
        "no_images": "No image files found in the selected folder.",
        "processing_n": "Processing {n} images…",
        "done_saved_to": "Done. Saved to:",
        "duplicate_in_session": "This image is already in the session.",
    },
    "id": {
        "window_title": "Deteksi Parasit Malaria",
        "left_panel": "Asli",
        "right_panel": "Hasil Deteksi",
        "counts_title_img": "Penghitungan — Gambar Ini",
        "counts_title_sess": "Penghitungan — Total Sesi",
        "hdr_totals": "Total keseluruhan",
        # Sidebar / menus
        "sidebar_title": "Kontrol",
        "menu_file": "Berkas",
        "menu_session": "Sesi",
        "menu_view": "Tampilan",
        "menu_choose": "Pilih…",
        "open_image": "Buka Gambar",
        "process_folder": "Proses Folder",
        "save_result": "Simpan Hasil (gambar ini)",
        "save_session": "Simpan Sesi (semua)",
        "add_to_session": "Tambah ke Sesi",
        "reset_session": "Reset Sesi",
        "fit": "Pas ke Jendela",
        "reset": "Reset Tampilan",
        "appearance_label": "Tampilan",
        "appearance_system": "Sistem",
        "appearance_light": "Terang",
        "appearance_dark": "Gelap",
        "lang_label": "Bahasa",
        "session_images": "Jumlah gambar sesi:",
        "sidebar_toggle_tip": "Tutup/Buka Sidebar",
        "sidebar_expand": "Buka",
        "sidebar_collapse": "Tutup",
        # Status / dialogs
        "status_idle": "Buka gambar untuk mulai. Scroll untuk zoom, drag untuk geser.",
        "loading_model": "Memuat model…",
        "model_loaded": "Model dimuat.",
        "weights_not_found_title": "Model tidak ditemukan",
        "weights_not_found_msg": "best.onnx tidak ditemukan di paket.",
        "choose_image_title": "Pilih gambar",
        "open_error_title": "Gagal Membuka",
        "open_error_msg": "Gagal membaca gambar.",
        "no_image_title": "Belum Ada Gambar",
        "no_image_msg": "Silakan buka gambar terlebih dahulu.",
        "running": "Menjalankan inferensi…",
        "no_results": "Tidak ada deteksi.",
        "detections_saved": "Deteksi disimpan:",
        "save_result_title": "Simpan Hasil",
        "png_filter": "PNG",
        "jpeg_filter": "JPEG",
        "all_files": "Semua berkas",
        "saved_prefix": "Tersimpan:",
        "save_error_title": "Gagal Menyimpan",
        "save_error_msg": "Gagal menulis gambar.",
        "loaded_prefix": "Dimuat:",
        "nothing_to_save": "Belum ada yang disimpan.\nJalankan deteksi lalu tambahkan ke Sesi.",
        "save_session_title": "Pilih Folder Sesi",
        "session_saved": "Sesi tersimpan di:",
        "add_before_detect": "Jalankan deteksi dahulu sebelum menambah ke sesi.",
        "added_entry": "Ditambahkan ke sesi.",
        "reset_done": "Sesi di-reset.",
        "run_detection": "Jalankan Deteksi",
        "step2_hint": "",
        "no_images": "Tidak ada berkas gambar di folder yang dipilih.",
        "processing_n": "Memproses {n} gambar…",
        "done_saved_to": "Selesai. Tersimpan di:",
        "duplicate_in_session": "Gambar ini sudah ada di sesi.",
    },
}
LANG_LABELS = {"en": "English", "id": "Indonesia"}

# ────────────────────────── Model / classes ──────────────────────────
CLASS_NAMES = [
    "Malariae_Gametocyte","Malariae_Trophozoite","Malariae_Schizont","Malariae_Ring",
    "Ovale_Gametocyte","Ovale_Ring","Ovale_Trophozoite",
    "Falciparum_Ring","Falciparum_Schizont","Falciparum_Trophozoite","Falciparum_Gametocyte",
    "Vivax_Gametocyte","Vivax_Ring","Vivax_Schizont","Vivax_Trophozoite",
]
NUM_CLASSES = len(CLASS_NAMES)
SPECIES = ["Malariae", "Ovale", "Falciparum", "Vivax"]
STAGES  = ["Ring", "Trophozoite", "Schizont", "Gametocyte"]
CLASS_TO_GRID = {ci: (SPECIES.index(c.split("_",1)[0]), STAGES.index(c.split("_",1)[1]))
                 for ci, c in enumerate(CLASS_NAMES)}

def counts_to_matrix(count_vec: np.ndarray) -> np.ndarray:
    mat = np.zeros((len(SPECIES), len(STAGES)), dtype=int)
    for ci, cnt in enumerate(count_vec):
        si, ti = CLASS_TO_GRID[ci]
        mat[si, ti] += int(cnt)
    return mat

IMGSZ = 640  # fixed

def resource_path(rel_path: str) -> str:
    base_path = getattr(sys, "_MEIPASS", None)
    return str((Path(base_path) if base_path else Path(__file__).parent) / rel_path)

def read_image_bgr(path: str):
    try:
        img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
        if img is not None: return img
    except Exception: pass
    try:
        pil = Image.open(path).convert("RGB")
        return cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2BGR)
    except Exception: return None

def bgr_to_lab(img_bgr: np.ndarray) -> np.ndarray:
    return cv2.cvtColor(img_bgr, cv2.COLOR_BGR2Lab)

def letterbox(img, new_shape=IMGSZ, color=(114,114,114)):
    h, w = img.shape[:2]
    r = min(new_shape / h, new_shape / w)
    nw, nh = int(round(w * r)), int(round(h * r))
    pw, ph = new_shape - nw, new_shape - nh
    l, rgt = int(pw // 2), int(pw - pw // 2)
    t, btm = int(ph // 2), int(ph - ph // 2)
    img = cv2.resize(img, (nw, nh), interpolation=cv2.INTER_LINEAR)
    img = cv2.copyMakeBorder(img, t, btm, l, rgt, cv2.BORDER_CONSTANT, value=color)
    return img, r, (l, t)

def box_iou(a, b):
    tl = np.maximum(a[:,None,:2], b[None,:,:2])
    br = np.minimum(a[:,None,2:], b[None,:,2:])
    wh = np.clip(br - tl, a_min=0, a_max=None)
    inter = wh[:,:,0]*wh[:,:,1]
    area_a = (a[:,2]-a[:,0])*(a[:,3]-a[:,1])
    area_b = (b[:,2]-b[:,0])*(b[:,3]-b[:,1])
    return inter / (area_a[:,None] + area_b[None,:] - inter + 1e-7)

def nms_np(boxes, scores, iou_thres=0.45):
    idxs = scores.argsort()[::-1]
    keep = []
    while idxs.size > 0:
        i = idxs[0]; keep.append(i)
        if idxs.size == 1: break
        ious = box_iou(boxes[i:i+1], boxes[idxs[1:]]).ravel()
        idxs = idxs[1:][ious <= iou_thres]
    return np.array(keep, dtype=np.int32)

# ────────────────────────── Predictor ──────────────────────────
class OnnxYoloV8:
    def __init__(self, onnx_path: str, num_classes: int):
        sess_opts = ort.SessionOptions()
        providers = ["CPUExecutionProvider"]
        self.session = ort.InferenceSession(onnx_path, sess_options=sess_opts, providers=providers)
        self.input_name = self.session.get_inputs()[0].name
        self.nc = num_classes
    def _preprocess(self, img_bgr):
        lab = bgr_to_lab(img_bgr)
        lb, ratio, (padw, padh) = letterbox(lab, new_shape=IMGSZ, color=(114,114,114))
        lb = lb[:,:,::-1].copy()
        inp = (lb.astype(np.float32)/255.0).transpose(2,0,1)[None,...]
        return inp, ratio, padw, padh
    def _postprocess(self, pred, ratio, padw, padh, orig_shape, conf_thres=0.25, iou_thres=0.45):
        if pred.ndim == 3: pred = np.squeeze(pred, 0)
        if pred.shape[0] == (4 + self.nc) and pred.shape[1] != (4 + self.nc): pred = pred.T
        boxes_xywh = pred[:, :4]; scores_all = pred[:, 4:]
        if scores_all.max() > 1.0 or scores_all.min() < 0.0:
            scores_all = 1.0 / (1.0 + np.exp(-scores_all))
        cls_idx = np.argmax(scores_all, axis=1)
        cls_scores = scores_all[np.arange(scores_all.shape[0]), cls_idx]
        keep = cls_scores >= conf_thres
        if not np.any(keep): return []
        boxes_xywh = boxes_xywh[keep]; cls_idx = cls_idx[keep]; cls_scores = cls_scores[keep]
        x,y,w,h = boxes_xywh[:,0], boxes_xywh[:,1], boxes_xywh[:,2], boxes_xywh[:,3]
        x = (x - padw)/ratio; y = (y - padh)/ratio; w = w/ratio; h = h/ratio
        x1 = np.clip(x - w/2, 0, orig_shape[1]-1)
        y1 = np.clip(y - h/2, 0, orig_shape[0]-1)
        x2 = np.clip(x + w/2, 0, orig_shape[1]-1)
        y2 = np.clip(y + h/2, 0, orig_shape[0]-1)
        boxes_xyxy = np.stack([x1,y1,x2,y2], axis=1)
        keep_idx = nms_np(boxes_xyxy, cls_scores, iou_thres=iou_thres)
        out = []
        for b, sc, c in zip(boxes_xyxy[keep_idx], cls_scores[keep_idx], cls_idx[keep_idx]):
            out.append([int(b[0]), int(b[1]), int(b[2]), int(b[3]), float(sc), int(c)])
        return out
    def predict(self, img_bgr, conf_thres=0.25, iou_thres=0.45):
        inp, ratio, padw, padh = self._preprocess(img_bgr)
        pred = self.session.run(None, {self.input_name: inp})[0]
        return self._postprocess(pred, ratio, padw, padh, img_bgr.shape, conf_thres, iou_thres)

# ──────────────────────── Drawing ────────────────────────
def draw_boxes_on_bgr(bgr_img: np.ndarray, dets, names):
    out = bgr_img.copy(); H,W = out.shape[:2]
    short = min(H,W); thickness=max(3,int(short/240)); font_scale=max(0.9,min(2.6,short/380))
    pad=max(4,thickness+2); line_gap=max(2,thickness-1); font=cv2.FONT_HERSHEY_SIMPLEX; text_thick=max(1,thickness-1)
    def text_size(s): (tw,th),base=cv2.getTextSize(s,font,font_scale,text_thick); return tw,th,base
    def wrap_text(label,max_w):
        if max_w<20: max_w=20
        words=label.replace("_"," ").split(" "); lines=[]; line=""
        for w in words:
            cand=(line+" "+w).strip() if line else w
            if text_size(cand)[0]<=max_w or not line: line=cand
            else: lines.append(line); line=w
        if line: lines.append(line)
        fixed=[]
        for ln in lines:
            if text_size(ln)[0]<=max_w: fixed.append(ln); continue
            buf=""
            for ch in ln:
                cand=buf+ch
                if text_size(cand)[0]<=max_w or not buf: buf=cand
                else: fixed.append(buf); buf=ch
            if buf: fixed.append(buf)
        return fixed
    def rect_overlaps(a,b):
        ax1,ay1,ax2,ay2=a; bx1,by1,bx2,by2=b
        return not (ax2<=bx1 or bx2<=ax1 or ay2<=by1 or by2<=ay1)
    def nearest_edge_point(px,py,rx1,ry1,rx2,ry2):
        x=np.clip(px,rx1,rx2); y=np.clip(py,ry1,ry2)
        inside=(rx1<x<rx2) and (ry1<y<ry2)
        if inside:
            d=[x-rx1, rx2-x, y-ry1, ry2-y]; i=int(np.argmin(d))
            if i==0: x=rx1
            elif i==1: x=rx2
            elif i==2: y=ry1
            else: y=ry2
        return int(x),int(y)
    GREEN=(0,220,120)
    for (x1,y1,x2,y2,*_) in dets:
        cv2.rectangle(out,(int(x1),int(y1)),(int(x2),int(y2)),GREEN,thickness)
    occupied=[]; dets_sorted=sorted(dets,key=lambda d:(int(d[1]),int(d[0])))
    for (x1,y1,x2,y2,conf,cls) in dets_sorted:
        x1,y1,x2,y2=map(int,(x1,y1,x2,y2))
        label=f"{names[cls] if 0<=cls<len(names) else cls} {conf:.2f}"
        def text_block(xc):
            xc=max(0,min(xc,W-1)); max_line_w=max(20,W-xc-2*pad)
            lines=wrap_text(label,max_line_w); sizes=[text_size(s) for s in lines]
            rw=max(s[0] for s in sizes)+2*pad
            rh=sum((s[1]+s[2]) for s in sizes)+(len(lines)-1)*line_gap+2*pad
            x0=min(max(0,xc),max(0,W-rw)); return x0,lines,sizes,rw,rh
        cands=[]; 
        for xa in (x1,x2):
            x0,lines,sizes,rw,rh=text_block(xa)
            cands+=[("above",x0,max(0,y1-rh),lines,sizes,rw,rh),("below",x0,min(H-rh,y2),lines,sizes,rw,rh)]
        x0,lines,sizes,rw,rh=text_block(x1)
        cands+=[("inside",x0,min(H-rh,y1),lines,sizes,rw,rh),("inside",x0,max(0,y2-rh),lines,sizes,rw,rh)]
        placed=None
        for _m,x0,y0,lines,sizes,rw,rh in cands:
            step=rh+line_gap; ok=False; y_try=y0
            for _ in range(20):
                rect=(x0,y_try,x0+rw,y_try+rh)
                if all(not rect_overlaps(rect,o) for o in occupied):
                    placed=(x0,y_try,lines,sizes,rw,rh); ok=True; break
                y_try=min(H-rh,y_try+step)
                if y_try==y0: break
            if not ok:
                y_try=y0
                for _ in range(20):
                    rect=(x0,y_try,x0+rw,y_try+rh)
                    if all(not rect_overlaps(rect,o) for o in occupied):
                        placed=(x0,y_try,lines,sizes,rw,rh); ok=True; break
                    y_try=max(0,y_try-step)
                    if y_try==y0: break
            if ok: break
        if placed is None:
            x0,lines,sizes,rw,rh=text_block(x1)
            placed=(x0,max(0,y1-rh),lines,sizes,rw,rh)
        x0,y0,lines,sizes,rw,rh=placed; occupied.append((x0,y0,x0+rw,y0+rh))
        cv2.rectangle(out,(x0,y0),(min(W-1,x0+rw),min(H-1,y0+rh)),(0,220,120),-1)
        y_cursor=y0+pad
        for (text,(tw,th,base)) in zip(lines,sizes):
            y_cursor+=th
            cv2.putText(out,text,(x0+pad,y_cursor),font,font_scale,(0,0,0),text_thick,cv2.LINE_AA)
            y_cursor+=base+line_gap
        bx_c=(x1+x2)/2.0; by_c=(y1+y2)/2.0; lx_c=x0+rw/2.0; ly_c=y0+rh/2.0
        sx,sy=nearest_edge_point(bx_c,by_c,x0,y0,x0+rw,y0+rh)
        ex,ey=nearest_edge_point(lx_c,ly_c,x1,y1,x2,y2)
        cv2.line(out,(sx,sy),(ex,ey),(0,0,0),text_thick+2,cv2.LINE_AA)
        cv2.line(out,(sx,sy),(ex,ey),(0,160,90),text_thick+1,cv2.LINE_AA)
        cv2.circle(out,(ex,ey),max(3,text_thick+1),(0,220,120),-1,cv2.LINE_AA)
    return out

# ──────────────────────── Viewer ────────────────────────
class ViewerCard(ctk.CTkFrame):
    def __init__(self, master, title, **kwargs):
        super().__init__(master, corner_radius=16, **kwargs)
        self.grid_rowconfigure(1, weight=1); self.grid_columnconfigure(0, weight=1)
        self.title = ctk.CTkLabel(self, text=title, font=ctk.CTkFont(size=16, weight="bold"))
        self.title.grid(row=0, column=0, sticky="w", padx=14, pady=(8,4))
        self.canvas = tk.Canvas(self, highlightthickness=0, bd=0)
        self.canvas.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0,8))
        self._img_pil=None; self._tk_img=None; self._scale=1.0; self._offset=[0.0,0.0]; self._drag_start=None
        self.canvas.bind("<Configure>", self._on_configure)
        self.canvas.bind("<ButtonPress-1>", self._on_press)
        self.canvas.bind("<B1-Motion>", self._on_drag)
        self.canvas.bind("<MouseWheel>", self._on_wheel)
        self.canvas.bind("<Button-4>", lambda e: self._zoom_at(1.1,(e.x,e.y)))
        self.canvas.bind("<Button-5>", lambda e: self._zoom_at(1/1.1,(e.x,e.y)))
    def _canvas_bg(self):
        return "#ffffff" if ctk.get_appearance_mode().lower()=="light" else "#0b0b0b"
    def set_title(self, text: str): self.title.configure(text=text)
    def set_image_bgr(self, img_bgr):
        if img_bgr is None:
            self._img_pil=None; self._tk_img=None; self.canvas.delete("all"); return
        img_rgb=cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
        self._img_pil=Image.fromarray(img_rgb); self.fit_to_window()
    def reset_view(self):
        if self._img_pil is None: return
        cw,ch=max(1,self.canvas.winfo_width()),max(1,self.canvas.winfo_height())
        iw,ih=self._img_pil.size
        self._scale=1.0 if iw<=cw and ih<=ch else min(cw/iw, ch/ih)
        self._offset=[(cw-iw*self._scale)/2.0, (ch-ih*self._scale)/2.0]
        self._redraw()
    def fit_to_window(self):
        if self._img_pil is None: self.canvas.delete("all"); return
        cw,ch=max(1,self.canvas.winfo_width()),max(1,self.canvas.winfo_height())
        iw,ih=self._img_pil.size
        self._scale=min(cw/iw, ch/ih, 1.0)
        self._offset=[(cw-iw*self._scale)/2.0, (ch-ih*self._scale)/2.0]
        self._redraw()
    def _on_configure(self,_): self._redraw()
    def _on_press(self,e): self._drag_start=(e.x,e.y)
    def _on_drag(self,e):
        if not self._drag_start: return
        dx,dy=e.x-self._drag_start[0], e.y-self._drag_start[1]
        self._drag_start=(e.x,e.y); self._offset[0]+=dx; self._offset[1]+=dy; self._redraw()
    def _on_wheel(self,e):
        if e.delta>0: self._zoom_at(1.1,(e.x,e.y))
        elif e.delta<0: self._zoom_at(1/1.1,(e.x,e.y))
    def _zoom_at(self,factor,center):
        if self._img_pil is None: return
        old=self._scale; new=max(0.1, min(8.0, old*factor))
        if abs(new-old)<1e-6: return
        cx,cy=center; iw,ih=self._img_pil.size; ox,oy=self._offset
        relx=max(0.0,min(1.0,(cx-ox)/(iw*old))); rely=max(0.0,min(1.0,(cy-oy)/(ih*old)))
        self._scale=new; self._offset[0]=cx-relx*iw*new; self._offset[1]=cy-rely*ih*new; self._redraw()
    def _redraw(self):
        self.canvas.delete("all"); self.canvas.configure(bg=self._canvas_bg())
        if self._img_pil is None: return
        iw,ih=self._img_pil.size; sw,sh=max(1,int(iw*self._scale)),max(1,int(ih*self._scale))
        disp=self._img_pil.resize((sw,sh), Image.LANCZOS)
        self._tk_img=ImageTk.PhotoImage(disp, master=self.canvas)
        self.canvas.create_image(int(self._offset[0]), int(self._offset[1]), image=self._tk_img, anchor="nw")

# ───────────────────────────── App (fixed/collapsible sidebar + menus) ─────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.session_active = False
        self.session_list_items = []   # untuk label di scroll list
        ctk.set_appearance_mode("light"); ctk.set_default_color_theme("dark-blue")
        sys_loc=(locale.getdefaultlocale()[0] or "").lower()
        self.lang_code="id" if "id" in sys_loc else "en"; self.t=lambda k: LANG[self.lang_code][k]
        self.title(self.t("window_title")); self.geometry("1480x980"); self.minsize(1220,860)

        self.predictor=None
        self.input_path: Path|None=None; self.input_bgr=None; self.result_bgr=None
        self._last_dets=None  # RAM only for single detection
        self.image_counts=np.zeros(NUM_CLASSES, dtype=int)
        self.session_counts=np.zeros(NUM_CLASSES, dtype=int)
        self.session_entries=[]  # list of {"path", "counts", "dets"}

        self.grid_columnconfigure(0, weight=0) # sidebar/handle
        self.grid_columnconfigure(1, weight=1) # content
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_content()
        self._apply_language()

        # Shortcuts
        self.bind("<Control-o>", lambda e: self._on_open_image())
        self.bind("<Control-f>", lambda e: self._on_process_folder())
        self.bind("<Return>",     lambda e: self._on_run())
    def _set_session_indicator(self, active: bool):
        is_light = ctk.get_appearance_mode().lower() == "light"
        if active:
            text = "Sesi AKTIF" if self.lang_code == "id" else "Session ACTIVE"
            bg   = "#dcfce7" if is_light else "#064e3b"
            fg   = "#166534" if is_light else "#a7f3d0"
        else:
            text = "Tidak dalam sesi" if self.lang_code == "id" else "No active session"
            bg   = "#fee2e2" if is_light else "#7f1d1d"
            fg   = "#991b1b" if is_light else "#fecaca"

        # gunakan satu label saja; width sudah fixed jadi area lama tertutup
        self.session_status_lbl.configure(text=text, fg_color=bg, text_color=fg)
        self.session_status_lbl.update_idletasks()


    def _on_save_choice(self, choice):
        c = choice.lower()
        if self.t("save_result").lower() in c:
            self._on_save()
        elif self.t("save_session").lower() in c:
            self._on_save_session()
        self.save_opt.set("Simpan…")

    def _on_reset_choice(self, choice):
        c = choice.lower()
        if self.t("reset").lower() in c:
            self._on_reset()
        elif self.t("reset_session").lower() in c:
            self._on_reset_session()
        self.reset_opt.set("Reset…")

    def _on_start_session(self):
        # mulai sesi baru (clear sebelumnya)
        self.session_active = True
        self._set_session_indicator(True)
        self.session_counts[:] = 0
        self.session_entries.clear()
        self._refresh_session_list()
        self._update_counts_labels()
        self.status.configure(text="Sesi dimulai. Hasil deteksi berikutnya akan otomatis ditambahkan.")
        if hasattr(self, "session_status_lbl"):
            self.session_status_lbl.configure(
                text=("Sesi AKTIF" if self.lang_code == "id" else "Session ACTIVE"),
                text_color="green"
            )

    # ---------- Theme palette for tables ----------
    def _palette(self):
        mode = ctk.get_appearance_mode().lower()
        if mode == "light":
            return {"text":"#111827","header_bg":"#e5e7eb","zebra1":"#f9fafb","zebra2":"#f3f4f6","border":"#d1d5db"}
        else:
            return {"text":"#e5e7eb","header_bg":"#1f2937","zebra1":"#0b1220","zebra2":"#0e1626","border":"#374151"}

    # ---------- Sidebar & handle ----------
    def _build_sidebar(self):
        self.sidebar_width = 280
        self.sidebar = ctk.CTkFrame(self, width=self.sidebar_width, corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsw")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_columnconfigure(0, weight=1)

        # header
        hdr = ctk.CTkFrame(self.sidebar, corner_radius=0)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_columnconfigure(0, weight=1)
        self.sidebar_title_lbl = ctk.CTkLabel(hdr, text=self.t("sidebar_title"),
                                            font=ctk.CTkFont(size=16, weight="bold"))
        self.sidebar_title_lbl.grid(row=0, column=0, sticky="w", padx=12, pady=10)

        # >>> NEW: session status indicator
        self.session_status_lbl = ctk.CTkLabel(
            self.sidebar,
            text=("Tidak dalam sesi" if self.lang_code == "id" else "No active session"),
            font=ctk.CTkFont(size=12, weight="bold"),
            fg_color="#fee2e2",     # bg awal (tidak transparan)
            text_color="#991b1b",
            corner_radius=8,
            padx=8, pady=4,
            width=220,              # << lebar tetap supaya sisa teks tertutup
            anchor="w"              # rata kiri
        )
        self.session_status_lbl.grid(row=1, column=0, sticky="w", padx=12, pady=(0,8))

        row = 2 
        # --- Berkas: tombol ---
        self.lbl_file = ctk.CTkLabel(self.sidebar, text=self.t("menu_file"),
                                    font=ctk.CTkFont(size=13, weight="bold"))
        self.lbl_file.grid(row=row, column=0, sticky="w", padx=12, pady=(6,2)); row += 1

        self.btn_open_img = ctk.CTkButton(self.sidebar, text=self.t("open_image"),
                                        command=self._on_open_image, height=36)
        self.btn_open_img.grid(row=row, column=0, padx=12, pady=(0,6), sticky="ew"); row += 1

        self.btn_open_folder = ctk.CTkButton(self.sidebar, text=self.t("process_folder"),
                                            command=self._on_process_folder, height=36)
        self.btn_open_folder.grid(row=row, column=0, padx=12, pady=(0,8), sticky="ew"); row += 1

        # --- Sesi: tombol mulai + dropdown simpan/reset ---
        self.lbl_session = ctk.CTkLabel(self.sidebar, text=self.t("menu_session"),
                                        font=ctk.CTkFont(size=13, weight="bold"))
        self.lbl_session.grid(row=row, column=0, sticky="w", padx=12, pady=(6,2)); row += 1

        self.btn_start_session = ctk.CTkButton(self.sidebar, text="Mulai Sesi",
                                            command=self._on_start_session, height=36)
        self.btn_start_session.grid(row=row, column=0, padx=12, pady=(0,6), sticky="ew"); row += 1

        # dropdown Simpan
        self.save_opt = ctk.CTkOptionMenu(self.sidebar, values=["Simpan…",
                                                                self.t("save_result"),
                                                                self.t("save_session")],
                                        command=self._on_save_choice)
        self.save_opt.set("Simpan…")
        self.save_opt.grid(row=row, column=0, padx=12, pady=(0,6), sticky="ew"); row += 1

        # dropdown Reset
        self.reset_opt = ctk.CTkOptionMenu(self.sidebar, values=["Reset…",
                                                                self.t("reset"),
                                                                self.t("reset_session")],
                                        command=self._on_reset_choice)
        self.reset_opt.set("Reset…")
        self.reset_opt.grid(row=row, column=0, padx=12, pady=(0,8), sticky="ew"); row += 1

        # --- Tampilan & Bahasa ---
        self.lbl_appearance = ctk.CTkLabel(self.sidebar, text=self.t("appearance_label"),
                                        font=ctk.CTkFont(size=12))
        self.lbl_appearance.grid(row=row, column=0, sticky="w", padx=12, pady=(8,2)); row += 1

        self.appearance_opt = ctk.CTkOptionMenu(self.sidebar,
            values=[self.t("appearance_system"), self.t("appearance_light"), self.t("appearance_dark")],
            command=self._on_theme_change)
        self.appearance_opt.set(self.t("appearance_light"))   # default terang
        self.appearance_opt.grid(row=row, column=0, padx=12, pady=(0,8), sticky="ew"); row += 1

        self.lbl_lang = ctk.CTkLabel(self.sidebar, text=self.t("lang_label"),
                                    font=ctk.CTkFont(size=12))
        self.lbl_lang.grid(row=row, column=0, sticky="w", padx=12, pady=(8,2)); row += 1

        self.lang_opt = ctk.CTkOptionMenu(self.sidebar, values=[LANG_LABELS["en"], LANG_LABELS["id"]],
                                        command=self._on_lang_change)
        self.lang_opt.set(LANG_LABELS[self.lang_code])
        self.lang_opt.grid(row=row, column=0, padx=12, pady=(0,8), sticky="ew"); row += 1

        # --- Info sesi + scroll list ---
        self.lbl_count = ctk.CTkLabel(self.sidebar, text=self.t("session_images"),
                                    font=ctk.CTkFont(size=12))
        self.lbl_count.grid(row=row, column=0, sticky="w", padx=12, pady=(6,2)); row += 1

        self.session_info = ctk.CTkLabel(self.sidebar, text="0")
        self.session_info.grid(row=row, column=0, sticky="w", padx=12); row += 1

        self.session_scroll = ctk.CTkScrollableFrame(self.sidebar, height=160, corner_radius=8)
        self.session_scroll.grid(row=row, column=0, sticky="nsew", padx=12, pady=(4,8)); row += 1
        self.sidebar.grid_rowconfigure(row, weight=1)  # spacer bawah

        # status text
        self.status = ctk.CTkLabel(self.sidebar, text=self.t("status_idle"), wraplength=240, justify="left")
        self.status.grid(row=row, column=0, sticky="we", padx=12, pady=(6,10))

        # handle untuk expand saat collapse
        self.handle = ctk.CTkFrame(self, width=16, corner_radius=0)
        self.handle.grid_propagate(False)
        self.handle_btn = ctk.CTkButton(self.handle, text="▶", width=16, command=self._expand_sidebar,
                                        fg_color="transparent")
        self.handle_btn.place(relx=0.5, rely=0.02, anchor="n")
        self.handle.grid_remove()


    def _collapse_sidebar(self):
        self.sidebar.grid_remove()
        self.handle.grid(row=0, column=0, sticky="nsw")
    def _expand_sidebar(self):
        self.handle.grid_remove()
        self.sidebar.grid(row=0, column=0, sticky="nsw")

    # ---------- Menus handlers ----------
    def _on_file_choice(self, choice):
        c = choice.lower()
        if self.t("open_image").lower() in c: self._on_open_image()
        elif self.t("process_folder").lower() in c: self._on_process_folder()
        elif self.t("save_result").lower() in c: self._on_save()
        self.file_opt.set(self.t("menu_choose"))
    def _on_session_choice(self, choice):
        c = choice.lower()
        if self.t("add_to_session").lower() in c: self._on_add_to_session()
        elif self.t("save_session").lower() in c: self._on_save_session()
        elif self.t("reset_session").lower() in c: self._on_reset_session()
        self.session_opt.set(self.t("menu_choose"))
    def _on_view_choice(self, choice):
        c = choice.lower()
        if self.t("fit").lower() in c: self._on_fit()
        elif self.t("reset").lower() in c: self._on_reset()
        self.view_opt.set(self.t("menu_choose"))

    # ---------- Content ----------
    def _build_content(self):
        self.content = ctk.CTkFrame(self, corner_radius=0)
        self.content.grid(row=0, column=1, sticky="nsew")
        self.content.grid_rowconfigure(1, weight=1)  # viewers grow
        self.content.grid_columnconfigure(0, weight=1)

        self._build_counts_row()
        self._build_images_row()

        # single big run button
        bar = ctk.CTkFrame(self.content, corner_radius=0)
        bar.grid(row=2, column=0, sticky="ew", padx=12, pady=(0,10))
        bar.grid_columnconfigure(0, weight=1)
        self.big_run = ctk.CTkButton(bar, text=self.t("run_detection"),
                                     height=52, corner_radius=12,
                                     fg_color="#3B82F6", hover_color="#2563EB", text_color="white",
                                     command=self._on_run, state="disabled")
        self.big_run.grid(row=0, column=0, sticky="ew", padx=0, pady=(8,4))

    def _build_counts_row(self):
        if hasattr(self, "counts_frame"): self.counts_frame.destroy()
        pal = self._palette()
        self.counts_frame = ctk.CTkFrame(self.content, corner_radius=0)
        self.counts_frame.grid(row=0, column=0, sticky="ew", padx=12, pady=(8,4))
        self.counts_frame.grid_columnconfigure(0, weight=1)
        self.counts_frame.grid_columnconfigure(1, weight=1)
        self.card_img_counts = self._build_counts_card(self.counts_frame, "counts_title_img", pal)
        self.card_img_counts.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        self.card_sess_counts = self._build_counts_card(self.counts_frame, "counts_title_sess", pal)
        self.card_sess_counts.grid(row=0, column=1, sticky="nsew", padx=(6,0))

    def _make_cell(self, parent, text="", pal=None):
        cell = ctk.CTkFrame(parent, corner_radius=6, border_width=1, border_color=pal["border"])
        cell.grid_propagate(False); cell.configure(width=64, height=26)
        lbl = ctk.CTkLabel(cell, text=text, anchor="e", font=ctk.CTkFont(size=12),
                           text_color=pal["text"])
        lbl.grid(row=0, column=0, sticky="nsew", padx=6, pady=1)
        return cell, lbl

    def _build_counts_card(self, parent, title_key, pal):
        card = ctk.CTkFrame(parent, corner_radius=10); card.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(card, text=self.t(title_key),
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=pal["text"]).grid(row=0, column=0, sticky="w", padx=10, pady=(8,4))
        table = ctk.CTkFrame(card, corner_radius=8); table.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0,10))
        for c in range(5): table.grid_columnconfigure(c, weight=1)
        # header
        h_corner = ctk.CTkFrame(table, corner_radius=6, fg_color=pal["header_bg"])
        h_corner.grid(row=0, column=0, padx=6, pady=(6,3), sticky="nsew")
        ctk.CTkLabel(h_corner, text="", text_color=pal["text"]).pack(padx=6, pady=3)
        for j, st in enumerate(STAGES, start=1):
            h = ctk.CTkFrame(table, corner_radius=6, fg_color=pal["header_bg"])
            h.grid(row=0, column=j, padx=6, pady=(6,3), sticky="nsew")
            ctk.CTkLabel(h, text=st, anchor="center",
                         font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=pal["text"]).pack(padx=6, pady=3)
        # body
        cells=[]
        for i, sp in enumerate(SPECIES, start=1):
            zebra = pal["zebra1"] if i%2==1 else pal["zebra2"]
            spf = ctk.CTkFrame(table, corner_radius=6, fg_color=zebra)
            spf.grid(row=i, column=0, padx=(8,4), pady=1, sticky="nsew")
            ctk.CTkLabel(spf, text=sp, anchor="w", text_color=pal["text"],
                         font=ctk.CTkFont(size=12)).pack(padx=8, pady=3)
            row=[]
            for j in range(4):
                c,l = self._make_cell(table, "0", pal=pal)
                c.grid(row=i, column=j+1, padx=(4,8), pady=1, sticky="nsew")
                row.append(l)
            cells.append(row)
        card._cells = cells
        return card

    def _build_images_row(self):
        if hasattr(self, "images_frame"): self.images_frame.destroy()
        self.images_frame = ctk.CTkFrame(self.content, corner_radius=0)
        self.images_frame.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0,6))
        self.images_frame.grid_rowconfigure(0, weight=1)
        self.images_frame.grid_columnconfigure(0, weight=1)
        self.images_frame.grid_columnconfigure(1, weight=1)
        self.card_left  = ViewerCard(self.images_frame, title=self.t("left_panel"))
        self.card_right = ViewerCard(self.images_frame, title=self.t("right_panel"))
        self.card_left.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        self.card_right.grid(row=0, column=1, sticky="nsew", padx=(6,0))

    # ---------- Language / theme ----------
    def _on_lang_change(self, choice: str):
        c = choice.strip().lower()
        self.lang_code = "en" if c.startswith("eng") else "id"
        self._apply_language()
    def _apply_language(self):
    # rebind translator
        self.t = lambda k: LANG[self.lang_code][k]
        self.title(self.t("window_title"))
        self._set_session_indicator(self.session_active)
        # session indicator text (guarded)
        if hasattr(self, "session_status_lbl"):
            self._set_session_indicator(getattr(self, "session_active", False))
            if getattr(self, "session_active", False):
                self.session_status_lbl.configure(
                    text=("Sesi AKTIF" if self.lang_code == "id" else "Session ACTIVE"),
                    text_color="green"
                )
            else:
                self.session_status_lbl.configure(
                    text=("Tidak dalam sesi" if self.lang_code == "id" else "No active session"),
                    text_color="red"
                )
        # ===== Sidebar texts =====
        self.sidebar_title_lbl.configure(text=self.t("sidebar_title"))
        # indikator sesi
        self.lbl_file.configure(text=self.t("menu_file"))
        self.lbl_session.configure(text=self.t("menu_session"))
        self.lbl_appearance.configure(text=self.t("appearance_label"))
        self.lbl_lang.configure(text=self.t("lang_label"))
        self.lbl_count.configure(text=self.t("session_images"))

        # buttons
        self.btn_open_img.configure(text=self.t("open_image"))
        self.btn_open_folder.configure(text=self.t("process_folder"))
        # label untuk tombol mulai sesi — bisa diterjemahkan kalau mau
        self.btn_start_session.configure(text="Mulai Sesi" if self.lang_code == "id" else "Start Session")

        # dropdowns (save / reset)
        self.save_opt.configure(values=[
            "Simpan…" if self.lang_code == "id" else "Save…",
            self.t("save_result"),
            self.t("save_session"),
        ])
        self.save_opt.set("Simpan…" if self.lang_code == "id" else "Save…")

        self.reset_opt.configure(values=[
            "Reset…",
            self.t("reset"),
            self.t("reset_session"),
        ])
        self.reset_opt.set("Reset…")

        # appearance + language dropdown labels
        self.appearance_opt.configure(values=[
            self.t("appearance_system"),
            self.t("appearance_light"),
            self.t("appearance_dark"),
        ])
        self.lang_opt.configure(values=[LANG_LABELS["en"], LANG_LABELS["id"]])

        # session count + status
        self.session_info.configure(text=str(len(self.session_entries)))
        self.status.configure(text=self.t("status_idle"))

        # ===== Main content texts =====
        self.big_run.configure(text=self.t("run_detection"))
        self.card_left.set_title(self.t("left_panel"))
        self.card_right.set_title(self.t("right_panel"))

        # Rebuild counts UI to apply localized headers & light/dark palette
        self._build_counts_row()


    def _on_theme_change(self, choice: str):
        ch = choice.lower()
        if ch.startswith(LANG[self.lang_code]["appearance_system"].lower()[0:3]): mode = "system"
        elif ch.startswith(LANG[self.lang_code]["appearance_light"].lower()[0:3]): mode = "light"
        else: mode = "dark"
        ctk.set_appearance_mode(mode)
        self._set_session_indicator(self.session_active)
        self._build_counts_row(); self.card_left._redraw(); self.card_right._redraw()

    # ---------- Model ----------
    def _ensure_model(self):
        if self.predictor: return True
        self.status.configure(text=self.t("loading_model")); self.update_idletasks()
        onnx_path = resource_path("best.onnx")
        if not Path(onnx_path).exists():
            alt = (Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent)/"best.onnx"
            if alt.exists(): onnx_path = str(alt)
            else:
                messagebox.showerror(self.t("weights_not_found_title"), self.t("weights_not_found_msg"))
                return False
        self.predictor = OnnxYoloV8(onnx_path, NUM_CLASSES)
        self.status.configure(text=self.t("model_loaded"))
        return True

    # ---------- Helpers ----------
    # ── Progress popup helpers ──────────────────────────────────────────
    def _progress_open(self, total: int):
        # modal toplevel
        self._prog_total = max(1, int(total))
        self._prog_cancel = False

        self.prog_win = ctk.CTkToplevel(self)
        self.prog_win.title("Processing…")
        self.prog_win.geometry("520x220")
        self.prog_win.grab_set()               # modal
        self.prog_win.attributes("-topmost", True)
        self.prog_win.protocol("WM_DELETE_WINDOW", lambda: None)  # disable close while running

        ctk.CTkLabel(self.prog_win, text="Processing images…",
                    font=ctk.CTkFont(size=18, weight="bold")).pack(pady=(18, 6))

        self.prog_msg = ctk.CTkLabel(self.prog_win, text="")
        self.prog_msg.pack(pady=(0, 10))

        self.prog_bar = ctk.CTkProgressBar(self.prog_win, height=22)
        self.prog_bar.pack(fill="x", padx=24, pady=6)
        self.prog_bar.set(0.0)

        self.prog_pct = ctk.CTkLabel(self.prog_win, text="0% (0 / {})".format(self._prog_total))
        self.prog_pct.pack(pady=(0, 8))

        btn_row = ctk.CTkFrame(self.prog_win); btn_row.pack(fill="x", padx=18, pady=(8,14))
        btn_row.grid_columnconfigure(0, weight=1)
        self.prog_cancel_btn = ctk.CTkButton(btn_row, text="Cancel",
                                            fg_color="#ef4444", hover_color="#dc2626",
                                            command=lambda: setattr(self, "_prog_cancel", True))
        self.prog_cancel_btn.grid(row=0, column=0, sticky="e")

        self.prog_win.update_idletasks()

    def _progress_update(self, i: int, filename: str | None = None):
        # i starts from 0..total
        frac = min(1.0, max(0.0, i / float(self._prog_total)))
        self.prog_bar.set(frac)
        if filename:
            self.prog_msg.configure(text=filename)
        self.prog_pct.configure(text=f"{int(frac*100)}% ({i} / {self._prog_total})")
        # make sure UI paints
        self.prog_win.update_idletasks()

    def _progress_close(self):
        try:
            self.prog_win.grab_release()
            self.prog_win.destroy()
        except Exception:
            pass

    def _refresh_session_list(self):
        # hapus isi lama
        for w in getattr(self, "session_list_items", []):
            try: w.destroy()
            except: pass
        self.session_list_items = []
        # isi ulang
        for idx, e in enumerate(self.session_entries, start=1):
            txt = f"{idx:02d} • {e['path'].name}"
            lbl = ctk.CTkLabel(self.session_scroll, text=txt, anchor="w")
            lbl.pack(fill="x", padx=6, pady=2)
            self.session_list_items.append(lbl)
        # update counter
        self.session_info.configure(text=str(len(self.session_entries)))

    def _add_current_to_session(self, *, auto: bool):
        """Tambahkan gambar yang sedang aktif ke sesi. auto=True saat dipanggil dari _on_run."""
        if self.input_path is None or self._last_dets is None:
            return
        # anti duplikat berdasarkan path
        if any(e["path"] == self.input_path for e in self.session_entries):
            if not auto:
                # opsional tampilkan info kalau manual klik
                try:
                    messagebox.showinfo("Info", self.t("duplicate_in_session"))
                except Exception:
                    pass
            return
        counts = self._counts_from_dets(self._last_dets)
        self.session_entries.append({"path": self.input_path, "counts": counts.copy(), "dets": self._last_dets})
        self.session_counts += counts
        self._update_counts_labels()
        self._refresh_session_list()
        # status ringkas
        self.status.configure(text=f"Ditambah ke sesi: {self.input_path.name}")

    def _refresh_session_list(self):
    # hapus isi lama
        for w in getattr(self, "session_list_items", []):
            try: w.destroy()
            except: pass
        self.session_list_items = []
        for idx, e in enumerate(self.session_entries, start=1):
            txt = f"{idx:02d} • {e['path'].name}"
            lbl = ctk.CTkLabel(self.session_scroll, text=txt, anchor="w")
            lbl.pack(fill="x", padx=6, pady=2)
            self.session_list_items.append(lbl)
        self.session_info.configure(text=str(len(self.session_entries)))

    def _update_counts_labels(self):
        img_mat  = counts_to_matrix(self.image_counts)
        sess_mat = counts_to_matrix(self.session_counts)
        for si in range(4):
            for ti in range(4):
                self.card_img_counts._cells[si][ti].configure(text=str(int(img_mat[si, ti])))
        for si in range(4):
            for ti in range(4):
                self.card_sess_counts._cells[si][ti].configure(text=str(int(sess_mat[si, ti])))
        self.session_info.configure(text=str(len(self.session_entries)))

    def _counts_from_dets(self, dets):
        arr = np.zeros(NUM_CLASSES, dtype=int)
        for *_xyxy, _conf, cls in dets:
            if 0 <= int(cls) < NUM_CLASSES:
                arr[int(cls)] += 1
        return arr

    def _safe_stem(self, p: Path):
        s = p.stem.strip().replace(" ", "_")
        return "".join(ch for ch in s if ch.isalnum() or ch in ("_", "-", "."))[:100] or "image"

    # ---------- Actions ----------
    def _on_open_image(self):
        p = filedialog.askopenfilename(title=self.t("choose_image_title"),
            filetypes=[("Images","*.jpg *.jpeg *.png *.bmp *.tif *.tiff")])
        if not p: return
        img = read_image_bgr(p)
        if img is None:
            messagebox.showerror(self.t("open_error_title"), self.t("open_error_msg"))
            return
        self.input_path = Path(p); self.input_bgr = img
        self.result_bgr = None; self._last_dets = None
        self.image_counts[:] = 0
        self.card_left.set_image_bgr(img)
        self.card_right.set_image_bgr(None)
        self.status.configure(text=f"{self.t('loaded_prefix')} {self.input_path.name}")
        self.big_run.configure(state="normal")
        self._update_counts_labels()

    def _on_run(self):
        if self.input_bgr is None:
            messagebox.showwarning(self.t("no_image_title"), self.t("no_image_msg")); return
        if not self._ensure_model(): return
        self.status.configure(text=self.t("running")); self.update_idletasks()
        try:
            dets = self.predictor.predict(self.input_bgr, conf_thres=0.25, iou_thres=0.45)
        except Exception as e:
            messagebox.showerror("Error", f"Inference failed: {e}")
            return
        self._last_dets = dets  # keep in RAM only
        if len(dets) == 0:
            self.card_right.set_image_bgr(self.input_bgr.copy())
            self.status.configure(text=self.t("no_results"))
            self.image_counts[:] = 0
            self._update_counts_labels()
            return
        self.result_bgr = draw_boxes_on_bgr(self.input_bgr, dets, CLASS_NAMES)
        self.card_right.set_image_bgr(self.result_bgr)
        self.image_counts = self._counts_from_dets(dets)
        # otomatis masuk ke sesi bila sesi aktif
        if self.session_active and self.input_path is not None:
            if not any(e["path"] == self.input_path for e in self.session_entries):
                entry = {"path": self.input_path, "counts": self.image_counts.copy(), "dets": self._last_dets}
                self.session_entries.append(entry)
                self.session_counts += self.image_counts
                self._update_counts_labels()
                self._refresh_session_list()
        self._update_counts_labels()
        # --- AUTO ADD TO SESSION SAAT SESI AKTIF ---
        if getattr(self, "session_active", False):
            self._add_current_to_session(auto=True)
        # NOTE: Do NOT write TXT or image copy here (RAM only)

    def _on_save(self):
        if self.result_bgr is None:
            messagebox.showwarning(self.t("no_image_title"), self.t("no_image_msg")); return
        default_name = (self.input_path.with_stem(self.input_path.stem + "_pred")
                        .with_suffix(".png").name if self.input_path else "result.png")
        p = filedialog.asksaveasfilename(title=self.t("save_result_title"),
            defaultextension=".png", initialfile=default_name,
            filetypes=[(self.t("png_filter"), "*.png"),
                       (self.t("jpeg_filter"), "*.jpg;*.jpeg"),
                       (self.t("all_files"), "*.*")])
        if not p: return
        ok = cv2.imwrite(p, self.result_bgr)
        if ok: self.status.configure(text=f"{self.t('saved_prefix')} {Path(p).name}")
        else:  messagebox.showerror(self.t("save_error_title"), self.t("save_error_msg"))

    def _on_process_folder(self):
        folder = filedialog.askdirectory(title=self.t("process_folder"))
        if not folder:
            return

        # Show popup immediately
        if not hasattr(self, "_progress_open"):
            messagebox.showerror("Missing helpers", "Progress helpers (_progress_open/_progress_update/_progress_close) not found.")
            return
        self._progress_open(1)
        self._progress_update(0, filename=self.t("loading_model"))

        # Ensure model (UI stays responsive because we already opened popup)
        if not self._ensure_model():
            self._progress_close()
            return

        # Scan files
        self._progress_update(0, filename=("Memindai berkas…" if self.lang_code == "id" else "Scanning files…"))
        exts = ("*.jpg", "*.jpeg", "*.png", "*.bmp", "*.tif", "*.tiff")
        files = []
        for e in exts:
            files.extend(Path(folder).glob(e))
        files = sorted(files)
        if not files:
            self._progress_close()
            messagebox.showinfo("No Images", self.t("no_images"))
            return

        # Prepare outputs
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = Path(folder) / f"batch_{ts}"
        det_dir = out_dir / "detections"
        det_dir.mkdir(parents=True, exist_ok=True)

        # Switch popup to real total and status
        self._prog_total = len(files)
        self.prog_bar.set(0.0)
        self.prog_pct.configure(text=f"0% (0 / {self._prog_total})")
        self._progress_update(0, filename=("Menyiapkan output…" if self.lang_code == "id" else "Preparing outputs…"))

        self.status.configure(text=self.t("processing_n").format(n=len(files)))

        # Initialize batch context (stored on self so worker can access)
        self._batch_ctx = {
            "files": files,
            "folder": Path(folder),
            "out_dir": out_dir,
            "det_dir": det_dir,
            "ts": ts,
            "i": 0,
            "batch_total_counts": np.zeros(NUM_CLASSES, dtype=int),
            "per_image_rows": [],
            "all_dets_rows": []
        }

        # Start worker on next loop tick (lets popup fully render)
        self.after(10, self._process_folder_worker)
    def _process_folder_worker(self):
        ctx = getattr(self, "_batch_ctx", None)
        if ctx is None:
            # nothing to do
            try: self._progress_close()
            except: pass
            return

        files = ctx["files"]
        i = ctx["i"]

        # Done or cancelled?
        if i >= len(files) or getattr(self, "_prog_cancel", False):
            # Write summaries
            out_dir = ctx["out_dir"]; det_dir = ctx["det_dir"]; ts = ctx["ts"]
            batch_total_counts = ctx["batch_total_counts"]
            per_image_rows = ctx["per_image_rows"]; all_dets_rows = ctx["all_dets_rows"]

            try:
                with open(out_dir / "summary_total.csv", "w", newline="", encoding="utf-8") as f:
                    w = csv.writer(f); w.writerow(["class_name","count"])
                    for ci, name in enumerate(CLASS_NAMES): w.writerow([name, int(batch_total_counts[ci])])

                with open(out_dir / "per_image.csv", "w", newline="", encoding="utf-8") as f:
                    fieldnames = ["image_index","image_name"] + CLASS_NAMES
                    w = csv.DictWriter(f, fieldnames=fieldnames); w.writeheader()
                    for row in per_image_rows: w.writerow(row)

                with open(out_dir / "summary.txt", "w", encoding="utf-8") as f:
                    f.write("Malaria Parasite Detection — Batch Summary\n")
                    f.write(f"Timestamp: {ts}\n")
                    f.write(f"Total images: {len(per_image_rows)}\n\n")
                    f.write(f"{self.t('hdr_totals')}:\n")
                    for ci, name in enumerate(CLASS_NAMES):
                        f.write(f"  {name}: {int(batch_total_counts[ci])}\n")
                    f.write("\nPer image recap:\n")
                    for row in per_image_rows:
                        f.write(f"- Image {row['image_index']}: {row['image_name']}\n")
                        for name in CLASS_NAMES:
                            cnt = row[name]
                            if cnt > 0: f.write(f"    {name}: {cnt}\n")
                        f.write("\n")
                    f.write("Annotated images & detection TXT are in the 'detections' folder.\n")

                # Excel (optional)
                if Workbook is not None:
                    from openpyxl import Workbook as _WB
                    from openpyxl.utils import get_column_letter
                    xlsx_path = out_dir / "batch.xlsx"
                    wb = _WB()

                    ws1 = wb.active; ws1.title = "Summary_Total"
                    ws1.append(["class_name","count"])
                    for ci, name in enumerate(CLASS_NAMES):
                        ws1.append([name, int(batch_total_counts[ci])])

                    ws2 = wb.create_sheet("Per_Image")
                    ws2.append(["image_index","image_name"] + CLASS_NAMES)
                    for row in per_image_rows:
                        ws2.append([row["image_index"], row["image_name"]] + [row[n] for n in CLASS_NAMES])

                    ws3 = wb.create_sheet("Detections")
                    ws3.append(["image_index","image_name","class_name","conf","x1","y1","x2","y2"])
                    for r in all_dets_rows:
                        ws3.append([r["image_index"], r["image_name"], r["class_name"], r["conf"],
                                    r["x1"], r["y1"], r["x2"], r["y2"]])

                    ws4 = wb.create_sheet("Summary")
                    r = 1
                    ws4.cell(row=r, column=1, value="Malaria Parasite Detection — Batch Summary"); r += 1
                    ws4.cell(row=r, column=1, value=f"Timestamp: {ts}"); r += 1
                    ws4.cell(row=r, column=1, value=f"Total images: {len(per_image_rows)}"); r += 2
                    ws4.cell(row=r, column=1, value=self.t("hdr_totals")); r += 1
                    ws4.cell(row=r, column=1, value="class_name"); ws4.cell(row=r, column=2, value="count"); r += 1
                    for ci, name in enumerate(CLASS_NAMES):
                        ws4.cell(row=r, column=1, value=name)
                        ws4.cell(row=r, column=2, value=int(batch_total_counts[ci])); r += 1
                    r += 1
                    ws4.cell(row=r, column=1, value="Per image recap:"); r += 1
                    for row in per_image_rows:
                        ws4.cell(row=r, column=1, value=f"Image {row['image_index']}: {row['image_name']}"); r += 1
                        for name in CLASS_NAMES:
                            cnt = row[name]
                            if cnt > 0:
                                ws4.cell(row=r, column=2, value=name)
                                ws4.cell(row=r, column=3, value=cnt); r += 1
                        r += 1

                    for ws in (ws1, ws2, ws3, ws4):
                        for col in ws.columns:
                            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                            ws.column_dimensions[get_column_letter(col[0].column)].width = min(50, max(10, max_len + 2))
                    wb.save(str(xlsx_path))

            except Exception as e:
                print(f"Batch summary failed: {e}")

            # Close popup and finalize
            try: self._progress_close()
            except: pass
            self.status.configure(text=f"{self.t('done_saved_to')} {ctx['out_dir'].name}")
            try:
                if os.name == 'nt':
                    os.startfile(str(ctx["out_dir"]))
                elif sys.platform == 'darwin':
                    subprocess.Popen(['open', str(ctx["out_dir"])])
                else:
                    subprocess.Popen(['xdg-open', str(ctx["out_dir"])])
            except Exception:
                pass

            # cleanup
            self._batch_ctx = None
            return

        # Not done: process one file
        img_path = files[i]
        self._progress_update(i, filename=img_path.name)

        img = read_image_bgr(str(img_path))
        if img is None:
            # advance to next
            ctx["i"] = i + 1
            self.after(1, self._process_folder_worker)
            return

        try:
            dets = self.predictor.predict(img, conf_thres=0.25, iou_thres=0.45)
        except Exception as e:
            print(f"Inference failed for {img_path}: {e}")
            dets = []

        # visualize & save into batch folder
        vis = draw_boxes_on_bgr(img, dets, CLASS_NAMES)
        try:
            cv2.imwrite(str(ctx["det_dir"] / f"{img_path.stem}_pred.png"), vis)
        except Exception as e:
            print(f"Save image failed for {img_path}: {e}")
        try:
            with open(ctx["det_dir"] / f"{img_path.stem}_pred.txt", "w", encoding="utf-8") as f:
                for (x1, y1, x2, y2, conf, cls) in dets:
                    cname = CLASS_NAMES[cls] if 0 <= cls < len(CLASS_NAMES) else str(cls)
                    f.write(f"{cname}\t{conf:.4f}\t{x1}\t{y1}\t{x2}\t{y2}\n")
        except Exception as e:
            print(f"Save txt failed for {img_path}: {e}")

        # accumulate
        counts = self._counts_from_dets(dets)
        ctx["batch_total_counts"] += counts
        row = {"image_index": i + 1, "image_name": Path(img_path).name}
        for ci, name in enumerate(CLASS_NAMES):
            row[name] = int(counts[ci])
        ctx["per_image_rows"].append(row)
        for (x1, y1, x2, y2, conf, cls) in dets:
            cname = CLASS_NAMES[cls] if 0 <= cls < len(CLASS_NAMES) else str(cls)
            ctx["all_dets_rows"].append({
                "image_index": i + 1, "image_name": Path(img_path).name,
                "class_name": cname, "conf": f"{conf:.4f}", "x1": x1, "y1": y1, "x2": x2, "y2": y2
            })

        # progress post-file
        self._progress_update(i + 1, filename=img_path.name)

        # schedule next
        ctx["i"] = i + 1
        self.after(1, self._process_folder_worker)

    def _on_add_to_session(self):
        if self.input_path is None or self._last_dets is None:
            messagebox.showwarning(self.t("no_image_title"), self.t("add_before_detect")); return
        # prevent duplicate by path
        if any(e["path"] == self.input_path for e in self.session_entries):
            messagebox.showinfo("Info", self.t("duplicate_in_session")); return
        counts = self._counts_from_dets(self._last_dets)
        self.session_counts += counts
        self.session_entries.append({"path": self.input_path, "counts": counts.copy(), "dets": self._last_dets})
        self._update_counts_labels()
        self.status.configure(text=self.t("added_entry"))

    def _on_reset_session(self):
        self.session_active = False
        self.session_counts[:] = 0
        self.session_entries.clear()
        self._set_session_indicator(False)
        self._refresh_session_list()
        self._update_counts_labels()
        self.status.configure(text=self.t("reset_done"))
        if hasattr(self, "session_status_lbl"):
            self.session_status_lbl.configure(
                text=("Tidak dalam sesi" if self.lang_code == "id" else "No active session"),
                text_color="red"
            )

    def _on_save_session(self):
        if len(self.session_entries) == 0:
            messagebox.showinfo(self.t("save_session_title"), self.t("nothing_to_save")); return
        out_dir = filedialog.askdirectory(title=self.t("save_session_title"))
        if not out_dir: return
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        sess_dir = Path(out_dir) / f"session_{ts}"
        det_dir = sess_dir / "detections"
        det_dir.mkdir(parents=True, exist_ok=True)

        per_image_rows = []; all_dets_rows = []
        for idx, e in enumerate(self.session_entries, start=1):
            img_path: Path = e["path"]; dets = e["dets"]
            img = read_image_bgr(str(img_path))
            if img is None: continue
            vis = draw_boxes_on_bgr(img, dets, CLASS_NAMES)
            out_img = det_dir / f"{idx:03d}_{self._safe_stem(img_path)}_pred.png"
            cv2.imwrite(str(out_img), vis)
            row = {"image_index": idx, "image_name": img_path.name}
            for i, name in enumerate(CLASS_NAMES): row[name] = int(e["counts"][i])
            per_image_rows.append(row)
            out_txt = det_dir / f"{idx:03d}_{self._safe_stem(img_path)}_pred.txt"
            with open(out_txt, "w", encoding="utf-8") as f:
                for (x1,y1,x2,y2,conf,cls) in dets:
                    cname = CLASS_NAMES[cls] if 0<=cls<len(CLASS_NAMES) else str(cls)
                    f.write(f"{cname}\t{conf:.4f}\t{x1}\t{y1}\t{x2}\t{y2}\n")
                    all_dets_rows.append({"image_index": idx,"image_name": img_path.name,
                                          "class_name": cname,"conf": f"{conf:.4f}",
                                          "x1": x1,"y1": y1,"x2": x2,"y2": y2})

        with open(sess_dir / "summary_total.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["class_name","count"])
            for i, name in enumerate(CLASS_NAMES): w.writerow([name, int(self.session_counts[i])])

        with open(sess_dir / "per_image.csv", "w", newline="", encoding="utf-8") as f:
            fieldnames = ["image_index","image_name"] + CLASS_NAMES
            w = csv.DictWriter(f, fieldnames=fieldnames); w.writeheader()
            for row in per_image_rows: w.writerow(row)

        summary_txt = sess_dir / "summary.txt"
        with open(summary_txt, "w", encoding="utf-8") as f:
            f.write("Malaria Parasite Detection — Session Summary\n")
            f.write(f"Timestamp: {ts}\n")
            f.write(f"Total images: {len(self.session_entries)}\n\n")
            f.write(f"{self.t('hdr_totals')}:\n")
            for i, name in enumerate(CLASS_NAMES): f.write(f"  {name}: {int(self.session_counts[i])}\n")
            f.write("\nPer image recap:\n")
            for idx, e in enumerate(self.session_entries, start=1):
                f.write(f"- Image {idx}: {e['path'].name}\n")
                for i, name in enumerate(CLASS_NAMES):
                    cnt = int(e["counts"][i])
                    if cnt > 0: f.write(f"    {name}: {cnt}\n")
                f.write("\n")
            f.write("Annotated images & detection TXT are in the 'detections' folder.\n")

        if Workbook is not None:
            xlsx_path = sess_dir / "session.xlsx"
            try:
                wb = Workbook()
                ws1 = wb.active; ws1.title = "Summary_Total"
                ws1.append(["class_name","count"])
                for i, name in enumerate(CLASS_NAMES): ws1.append([name, int(self.session_counts[i])])

                ws2 = wb.create_sheet("Per_Image")
                ws2.append(["image_index","image_name"] + CLASS_NAMES)
                for row in per_image_rows:
                    ws2.append([row["image_index"], row["image_name"]] + [row[n] for n in CLASS_NAMES])

                ws3 = wb.create_sheet("Detections")
                ws3.append(["image_index","image_name","class_name","conf","x1","y1","x2","y2"])
                for r in all_dets_rows:
                    ws3.append([r["image_index"], r["image_name"], r["class_name"], r["conf"],
                                r["x1"], r["y1"], r["x2"], r["y2"]])

                ws4 = wb.create_sheet("Summary")
                r = 1
                ws4.cell(row=r, column=1, value="Malaria Parasite Detection — Session Summary"); r += 1
                ws4.cell(row=r, column=1, value=f"Timestamp: {ts}"); r += 1
                ws4.cell(row=r, column=1, value=f"Total images: {len(self.session_entries)}"); r += 2
                ws4.cell(row=r, column=1, value=self.t("hdr_totals")); r += 1
                ws4.cell(row=r, column=1, value="class_name"); ws4.cell(row=r, column=2, value="count"); r += 1
                for i, name in enumerate(CLASS_NAMES):
                    ws4.cell(row=r, column=1, value=name)
                    ws4.cell(row=r, column=2, value=int(self.session_counts[i])); r += 1
                r += 1
                ws4.cell(row=r, column=1, value="Per image recap:"); r += 1
                for idx, e in enumerate(self.session_entries, start=1):
                    ws4.cell(row=r, column=1, value=f"Image {idx}: {e['path'].name}"); r += 1
                    for i, name in enumerate(CLASS_NAMES):
                        cnt = int(e["counts"][i])
                        if cnt > 0:
                            ws4.cell(row=r, column=2, value=name)
                            ws4.cell(row=r, column=3, value=cnt); r += 1
                    r += 1
                for ws in (ws1, ws2, ws3, ws4):
                    for col in ws.columns:
                        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                        ws.column_dimensions[get_column_letter(col[0].column)].width = min(50, max(10, max_len + 2))
                wb.save(str(xlsx_path))
            except Exception as e:
                print(f"Excel creation failed: {e}")

        self.status.configure(text=f"{self.t('session_saved')} {sess_dir.name}")
        try:
            if os.name == 'nt': os.startfile(str(sess_dir))
            elif sys.platform == 'darwin': subprocess.Popen(['open', str(sess_dir)])
            else: subprocess.Popen(['xdg-open', str(sess_dir)])
        except: pass

    # View helpers
    def _on_fit(self): self.card_left.fit_to_window(); self.card_right.fit_to_window()
    def _on_reset(self): self.card_left.reset_view(); self.card_right.reset_view()

def main():
    app = App()
    app.mainloop()

if __name__ == "__main__":
    main()
